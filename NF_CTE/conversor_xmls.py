#!/usr/bin/env python3
"""
Conversor NF-e / CT-e  →  PDF (DANFE/DACTE) e/ou Excel
Interface gráfica com CustomTkinter

Dependências:
    pip install customtkinter brazilfiscalreport qrcode[pil] pillow openpyxl
"""

import io
import logging
import os
import re
import sys
import threading
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

# ── GUI ───────────────────────────────────────────────────────────────────────
try:
    import customtkinter as ctk
    from tkinter import filedialog, messagebox
    from PIL import Image, ImageTk
except ImportError as e:
    print(f"Erro de importação GUI: {e}")
    print("Instale: pip install customtkinter pillow")
    sys.exit(1)

# ── Fiscal ────────────────────────────────────────────────────────────────────
try:
    from brazilfiscalreport.danfe import Danfe, DanfeConfig, ReceiptPosition
    from brazilfiscalreport.dacte import Dacte, DacteConfig
    FISCAL_OK = True
except ImportError:
    FISCAL_OK = False

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
LOG_FILE   = SCRIPT_DIR / "conversor.log"

NS = {
    "nfe": "http://www.portalfiscal.inf.br/nfe",
    "cte": "http://www.portalfiscal.inf.br/cte",
}

VERDE       = "#00C48C"
VERDE_ESC   = "#009E72"
FUNDO       = "#0F1117"
FUNDO_CARD  = "#1A1D27"
FUNDO_INPUT = "#12151E"
BORDA       = "#2A2D3E"
TEXTO       = "#E8EAF0"
TEXTO_SUB   = "#6B7080"
ERRO        = "#FF5252"
AVISO       = "#FFB347"
AZUL        = "#4A9FFF"

# ─────────────────────────────────────────────────────────────────────────────
#  LOGGER DE ARQUIVO
# ─────────────────────────────────────────────────────────────────────────────

def _setup_logger() -> logging.Logger:
    logger = logging.getLogger("conversor_xml")
    logger.setLevel(logging.DEBUG)

    if not logger.handlers:
        fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter(
            fmt="%(asctime)s  [%(levelname)-7s]  %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        ))
        logger.addHandler(fh)

    return logger

logger = _setup_logger()


# ─────────────────────────────────────────────────────────────────────────────
#  UTILITÁRIOS XML
# ─────────────────────────────────────────────────────────────────────────────

def _ns(tag: str, ns: str) -> str:
    return f"{{{NS[ns]}}}{tag}"


def detectar_tipo(caminho: Path) -> str:
    try:
        tree = ET.parse(caminho)
        root = tree.getroot()
        raw  = root.tag
        ns   = raw.split("}")[0].lstrip("{") if "}" in raw else ""
        if NS["nfe"] in ns:
            return "nfe"
        if NS["cte"] in ns:
            return "cte"
        for el in root.iter():
            t = el.tag.split("}")[-1] if "}" in el.tag else el.tag
            if t == "infNFe":
                return "nfe"
            if t in ("infCte", "infCTe"):
                return "cte"
    except ET.ParseError:
        pass
    return "desconhecido"


def _find(root, *caminhos, ns_key="nfe"):
    """Busca um valor em múltiplos caminhos possíveis."""
    for path in caminhos:
        partes = path.split("/")
        el = root
        for p in partes:
            found = el.find(_ns(p, ns_key))
            if found is None:
                el = None
                break
            el = found
        if el is not None and el.text:
            return el.text.strip()
    return ""


def _find_all(root, path, ns_key="nfe"):
    partes = path.split("/")
    els = [root]
    for p in partes:
        next_els = []
        for e in els:
            next_els.extend(e.findall(_ns(p, ns_key)))
        els = next_els
    return els


def formatar_cnpj_cpf(v: str) -> str:
    v = re.sub(r"\D", "", v)
    if len(v) == 14:
        return f"{v[:2]}.{v[2:5]}.{v[5:8]}/{v[8:12]}-{v[12:]}"
    if len(v) == 11:
        return f"{v[:3]}.{v[3:6]}.{v[6:9]}-{v[9:]}"
    return v


def formatar_data(v: str) -> str:
    if not v:
        return ""
    try:
        return datetime.fromisoformat(v[:10]).strftime("%d/%m/%Y")
    except Exception:
        return v[:10]


def moeda(v: str) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return v


# ─────────────────────────────────────────────────────────────────────────────
#  EXTRAÇÃO NF-e
# ─────────────────────────────────────────────────────────────────────────────

def extrair_nfe(caminho: Path) -> dict:
    tree = ET.parse(caminho)
    root = tree.getroot()

    # Navega até infNFe
    def find_infnfe(el):
        t = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if t == "infNFe":
            return el
        for child in el:
            r = find_infnfe(child)
            if r is not None:
                return r
        return None

    inf = find_infnfe(root)
    if inf is None:
        return {}

    def f(*paths):
        return _find(inf, *paths, ns_key="nfe")

    chave = inf.get("Id", "").replace("NFe", "")

    # Produtos
    produtos = []
    for det in inf.findall(_ns("det", "nfe")):
        prod = det.find(_ns("prod", "nfe"))
        if prod is None:
            continue
        imposto = det.find(_ns("imposto", "nfe"))
        icms_val = ""
        ipi_val  = ""
        if imposto:
            for icms_grp in imposto.findall(_ns("ICMS", "nfe")):
                for child in icms_grp:
                    vv = child.find(_ns("vICMS", "nfe"))
                    if vv is not None and vv.text:
                        icms_val = vv.text
            ipi_grp = imposto.find(_ns("IPI", "nfe"))
            if ipi_grp is not None:
                for child in ipi_grp:
                    vv = child.find(_ns("vIPI", "nfe"))
                    if vv is not None and vv.text:
                        ipi_val = vv.text

        def fp(tag):
            el = prod.find(_ns(tag, "nfe"))
            return el.text.strip() if el is not None and el.text else ""

        produtos.append({
            "item":        det.get("nItem", ""),
            "codigo":      fp("cProd"),
            "descricao":   fp("xProd"),
            "ncm":         fp("NCM"),
            "cfop":        fp("CFOP"),
            "unidade":     fp("uCom"),
            "quantidade":  fp("qCom"),
            "vl_unitario": fp("vUnCom"),
            "vl_total":    fp("vProd"),
            "icms":        icms_val,
            "ipi":         ipi_val,
        })

    # Duplicatas
    dups = []
    cobr = inf.find(_ns("cobr", "nfe"))
    if cobr:
        for dup in cobr.findall(_ns("dup", "nfe")):
            def fd(tag):
                el = dup.find(_ns(tag, "nfe"))
                return el.text.strip() if el is not None and el.text else ""
            dups.append({
                "numero": fd("nDup"),
                "venc":   formatar_data(fd("dVenc")),
                "valor":  fd("vDup"),
            })

    return {
        "tipo":           "NF-e",
        "chave":          chave,
        "numero":         f("ide/nNF"),
        "serie":          f("ide/serie"),
        "data_emissao":   formatar_data(f("ide/dhEmi", "ide/dEmi")),
        "nat_operacao":   f("ide/natOp"),
        "tipo_nf":        "Saída" if f("ide/tpNF") == "1" else "Entrada",
        # Emitente
        "emit_cnpj":      formatar_cnpj_cpf(f("emit/CNPJ", "emit/CPF")),
        "emit_nome":      f("emit/xNome"),
        "emit_ie":        f("emit/IE"),
        "emit_uf":        f("emit/enderEmit/UF"),
        "emit_municipio": f("emit/enderEmit/xMun"),
        # Destinatário
        "dest_cnpj":      formatar_cnpj_cpf(f("dest/CNPJ", "dest/CPF")),
        "dest_nome":      f("dest/xNome"),
        "dest_ie":        f("dest/IE"),
        "dest_uf":        f("dest/enderDest/UF"),
        # Totais
        "vl_produtos":    f("total/ICMSTot/vProd"),
        "vl_frete":       f("total/ICMSTot/vFrete"),
        "vl_desconto":    f("total/ICMSTot/vDesc"),
        "vl_icms":        f("total/ICMSTot/vICMS"),
        "vl_ipi":         f("total/ICMSTot/vIPI"),
        "vl_pis":         f("total/ICMSTot/vPIS"),
        "vl_cofins":      f("total/ICMSTot/vCOFINS"),
        "vl_total_nf":    f("total/ICMSTot/vNF"),
        # Transporte
        "transp_modalidade": f("transp/modFrete"),
        "transp_nome":    f("transp/transporta/xNome"),
        # Info adicionais
        "inf_adicionais": f("infAdic/infCpl"),
        # Detalhes
        "produtos":       produtos,
        "duplicatas":     dups,
        "arquivo":        caminho.name,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  EXTRAÇÃO CT-e
# ─────────────────────────────────────────────────────────────────────────────

def extrair_cte(caminho: Path) -> dict:
    tree = ET.parse(caminho)
    root = tree.getroot()

    def find_infcte(el):
        t = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if t in ("infCte", "infCTe"):
            return el
        for child in el:
            r = find_infcte(child)
            if r is not None:
                return r
        return None

    inf = find_infcte(root)
    if inf is None:
        return {}

    def f(*paths):
        return _find(inf, *paths, ns_key="cte")

    chave = inf.get("Id", "").replace("CTe", "")

    return {
        "tipo":           "CT-e",
        "chave":          chave,
        "numero":         f("ide/nCT"),
        "serie":          f("ide/serie"),
        "data_emissao":   formatar_data(f("ide/dhEmi", "ide/dEmi")),
        "cfop":           f("ide/CFOP"),
        "nat_operacao":   f("ide/natOp", "ide/descServico"),
        # Emitente
        "emit_cnpj":      formatar_cnpj_cpf(f("emit/CNPJ")),
        "emit_nome":      f("emit/xNome"),
        "emit_ie":        f("emit/IE"),
        "emit_uf":        f("emit/enderEmit/UF"),
        # Remetente
        "rem_cnpj":       formatar_cnpj_cpf(f("rem/CNPJ", "rem/CPF")),
        "rem_nome":       f("rem/xNome"),
        # Destinatário
        "dest_cnpj":      formatar_cnpj_cpf(f("dest/CNPJ", "dest/CPF")),
        "dest_nome":      f("dest/xNome"),
        "dest_uf":        f("dest/enderDest/UF"),
        # Tomador
        "tom_cnpj":       formatar_cnpj_cpf(f("toma3/CNPJ", "toma4/CNPJ", "toma3/CPF")),
        "tom_nome":       f("toma3/xNome", "toma4/xNome"),
        # Valores
        "vl_total":       f("vPrest/vTPrest"),
        "vl_receber":     f("vPrest/vRec"),
        "vl_carga":       f("infCTeNorm/infCarga/vCarga", "infCarga/vCarga"),
        "modal":          f("ide/modal"),
        # Info adicionais
        "inf_adicionais": f("compl/xObs"),
        "arquivo":        caminho.name,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  GERADOR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def estilo_cabecalho(ws, row, cols, titulo, fill_color="1A6B4A"):
    fill   = PatternFill("solid", fgColor=fill_color)
    fonte  = Font(bold=True, color="FFFFFF", size=10)
    borda  = Border(
        bottom=Side(style="medium", color="FFFFFF"),
    )
    for col, label in zip(range(1, cols + 1), titulo):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font      = fonte
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda


def estilo_linha(ws, row, valores, zebra=False):
    fill_color = "F0FFF8" if zebra else "FFFFFF"
    fill       = PatternFill("solid", fgColor=fill_color)
    fonte      = Font(size=9)
    borda_lado = Side(style="thin", color="D0D0D0")
    borda      = Border(
        left=borda_lado, right=borda_lado,
        top=borda_lado, bottom=borda_lado,
    )
    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = fonte
        cell.fill      = fill
        cell.border    = borda
        cell.alignment = Alignment(vertical="center")


def auto_largura(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def gerar_excel(registros_nfe: list, registros_cte: list, caminho_saida: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrão

    VERDE_ESCURO = "0D5C36"
    AZUL_ESCURO  = "0D3A6B"
    ROXO_ESCURO  = "3D1A6B"

    # ── ABA NF-e ──────────────────────────────────────────────────────────────
    if registros_nfe:
        ws_nfe = wb.create_sheet("NF-e — Resumo")
        ws_nfe.row_dimensions[1].height = 30

        cab_nfe = [
            "Arquivo", "Chave NF-e", "Número", "Série", "Data Emissão",
            "Tipo", "Natureza Operação",
            "Emitente CNPJ", "Emitente Nome", "Emitente UF",
            "Destinatário CNPJ", "Destinatário Nome", "Destinatário UF",
            "Vl. Produtos", "Vl. Frete", "Vl. Desconto",
            "Vl. ICMS", "Vl. IPI", "Vl. PIS", "Vl. COFINS",
            "Vl. Total NF",
        ]
        estilo_cabecalho(ws_nfe, 1, len(cab_nfe), cab_nfe, VERDE_ESCURO)

        for i, r in enumerate(registros_nfe, 2):
            def n(k):
                try:
                    v = r.get(k, "")
                    return float(v) if v else 0.0
                except Exception:
                    return r.get(k, "")

            linha = [
                r.get("arquivo", ""), r.get("chave", ""),
                r.get("numero", ""), r.get("serie", ""), r.get("data_emissao", ""),
                r.get("tipo_nf", ""), r.get("nat_operacao", ""),
                r.get("emit_cnpj", ""), r.get("emit_nome", ""), r.get("emit_uf", ""),
                r.get("dest_cnpj", ""), r.get("dest_nome", ""), r.get("dest_uf", ""),
                n("vl_produtos"), n("vl_frete"), n("vl_desconto"),
                n("vl_icms"), n("vl_ipi"), n("vl_pis"), n("vl_cofins"),
                n("vl_total_nf"),
            ]
            estilo_linha(ws_nfe, i, linha, i % 2 == 0)

        # Formata colunas de valor como moeda
        for row in ws_nfe.iter_rows(min_row=2, min_col=14, max_col=21):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        auto_largura(ws_nfe)

        # ── ABA Produtos NF-e ─────────────────────────────────────────────────
        ws_prod = wb.create_sheet("NF-e — Produtos")
        ws_prod.row_dimensions[1].height = 30

        cab_prod = [
            "Arquivo", "NF Número", "Item", "Código", "Descrição",
            "NCM", "CFOP", "Unidade", "Quantidade",
            "Vl. Unitário", "Vl. Total", "ICMS", "IPI",
        ]
        estilo_cabecalho(ws_prod, 1, len(cab_prod), cab_prod, VERDE_ESCURO)

        row_p = 2
        for r in registros_nfe:
            for p in r.get("produtos", []):
                def pn(k):
                    try:
                        v = p.get(k, "")
                        return float(v) if v else 0.0
                    except Exception:
                        return p.get(k, "")

                linha = [
                    r.get("arquivo", ""), r.get("numero", ""),
                    p.get("item", ""), p.get("codigo", ""), p.get("descricao", ""),
                    p.get("ncm", ""), p.get("cfop", ""), p.get("unidade", ""),
                    pn("quantidade"), pn("vl_unitario"), pn("vl_total"),
                    pn("icms"), pn("ipi"),
                ]
                estilo_linha(ws_prod, row_p, linha, row_p % 2 == 0)
                row_p += 1

        for row in ws_prod.iter_rows(min_row=2, min_col=10, max_col=13):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        auto_largura(ws_prod)

        # ── ABA Duplicatas ────────────────────────────────────────────────────
        ws_dup = wb.create_sheet("NF-e — Duplicatas")
        ws_dup.row_dimensions[1].height = 30
        cab_dup = ["Arquivo", "NF Número", "Nº Duplicata", "Vencimento", "Valor"]
        estilo_cabecalho(ws_dup, 1, len(cab_dup), cab_dup, VERDE_ESCURO)

        row_d = 2
        for r in registros_nfe:
            for d in r.get("duplicatas", []):
                try:
                    val = float(d.get("valor", 0))
                except Exception:
                    val = d.get("valor", "")
                linha = [
                    r.get("arquivo", ""), r.get("numero", ""),
                    d.get("numero", ""), d.get("venc", ""), val,
                ]
                estilo_linha(ws_dup, row_d, linha, row_d % 2 == 0)
                row_d += 1

        for row in ws_dup.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        auto_largura(ws_dup)

    # ── ABA CT-e ──────────────────────────────────────────────────────────────
    if registros_cte:
        ws_cte = wb.create_sheet("CT-e — Resumo")
        ws_cte.row_dimensions[1].height = 30

        cab_cte = [
            "Arquivo", "Chave CT-e", "Número", "Série", "Data Emissão",
            "CFOP", "Modal",
            "Emitente CNPJ", "Emitente Nome", "Emitente UF",
            "Remetente CNPJ", "Remetente Nome",
            "Destinatário CNPJ", "Destinatário Nome", "Destinatário UF",
            "Tomador CNPJ", "Tomador Nome",
            "Vl. Total Prestação", "Vl. a Receber", "Vl. Carga",
        ]
        estilo_cabecalho(ws_cte, 1, len(cab_cte), cab_cte, AZUL_ESCURO)

        for i, r in enumerate(registros_cte, 2):
            def cn(k):
                try:
                    v = r.get(k, "")
                    return float(v) if v else 0.0
                except Exception:
                    return r.get(k, "")

            linha = [
                r.get("arquivo", ""), r.get("chave", ""),
                r.get("numero", ""), r.get("serie", ""), r.get("data_emissao", ""),
                r.get("cfop", ""), r.get("modal", ""),
                r.get("emit_cnpj", ""), r.get("emit_nome", ""), r.get("emit_uf", ""),
                r.get("rem_cnpj",  ""), r.get("rem_nome",  ""),
                r.get("dest_cnpj", ""), r.get("dest_nome", ""), r.get("dest_uf", ""),
                r.get("tom_cnpj",  ""), r.get("tom_nome",  ""),
                cn("vl_total"), cn("vl_receber"), cn("vl_carga"),
            ]
            estilo_linha(ws_cte, i, linha, i % 2 == 0)

        for row in ws_cte.iter_rows(min_row=2, min_col=18, max_col=20):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        auto_largura(ws_cte)

    wb.save(caminho_saida)


# ─────────────────────────────────────────────────────────────────────────────
#  CONVERSÃO PDF
# ─────────────────────────────────────────────────────────────────────────────

def converter_pdf(caminho_xml: Path, caminho_pdf: Path, tipo: str,
                  danfe_cfg: "DanfeConfig", dacte_cfg: "DacteConfig") -> bool:
    if not FISCAL_OK:
        return False
    try:
        with open(caminho_xml, "rb") as f:
            xml_bytes = f.read()

        if tipo == "nfe":
            doc = Danfe(xml=xml_bytes, config=danfe_cfg)
        else:
            doc = Dacte(xml=xml_bytes, config=dacte_cfg)

        pdf_bytes = doc.output(dest="S")
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode("latin-1")
        caminho_pdf.write_bytes(pdf_bytes)
        return True
    except Exception as e:
        raise RuntimeError(str(e))


# ─────────────────────────────────────────────────────────────────────────────
#  INTERFACE GRÁFICA
# ─────────────────────────────────────────────────────────────────────────────

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Conversor NF-e / CT-e")
        self.geometry("860x740")
        self.minsize(800, 680)
        self.configure(fg_color=FUNDO)

        # Estado
        self.pasta_entrada  = ctk.StringVar()
        self.pasta_saida    = ctk.StringVar()
        self.gerar_pdf      = ctk.BooleanVar(value=True)
        self.gerar_excel    = ctk.BooleanVar(value=False)
        self.recibo_pos     = ctk.StringVar(value="topo")    # topo / baixo
        self.sobrescrever   = ctk.BooleanVar(value=False)
        self.processando    = False

        self._build_ui()

    # ── Construção da UI ──────────────────────────────────────────────────────

    def _build_ui(self):
        # Scrollable container principal
        self.scroll = ctk.CTkScrollableFrame(
            self, fg_color=FUNDO, scrollbar_button_color=BORDA,
            scrollbar_button_hover_color=VERDE,
        )
        self.scroll.pack(fill="both", expand=True, padx=0, pady=0)
        self.scroll.grid_columnconfigure(0, weight=1)

        row = 0

        # ── Header ────────────────────────────────────────────────────────────
        header = ctk.CTkFrame(self.scroll, fg_color=FUNDO_CARD,
                              corner_radius=0, border_width=0)
        header.grid(row=row, column=0, sticky="ew", padx=0, pady=(0, 2))
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="⚡  Conversor NF-e / CT-e",
            font=ctk.CTkFont(family="Helvetica Neue", size=22, weight="bold"),
            text_color=TEXTO,
        ).grid(row=0, column=0, sticky="w", padx=28, pady=(18, 2))

        ctk.CTkLabel(
            header,
            text="Gere DANFE, DACTE e planilhas Excel a partir dos seus XMLs fiscais",
            font=ctk.CTkFont(size=12),
            text_color=TEXTO_SUB,
        ).grid(row=1, column=0, sticky="w", padx=28, pady=(0, 16))

        row += 1

        # ── Pastas ────────────────────────────────────────────────────────────
        row = self._card(row, "📂  Pastas",
                         self._build_pastas)

        # ── Formato de saída ──────────────────────────────────────────────────
        row = self._card(row, "📄  Formato de Saída",
                         self._build_formato)

        # ── Opções ────────────────────────────────────────────────────────────
        row = self._card(row, "⚙️   Opções",
                         self._build_opcoes)

        # ── Botão converter ───────────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(self.scroll, fg_color="transparent")
        btn_frame.grid(row=row, column=0, sticky="ew", padx=24, pady=(8, 4))
        btn_frame.grid_columnconfigure(0, weight=1)
        row += 1

        self.btn_converter = ctk.CTkButton(
            btn_frame,
            text="▶   Iniciar Conversão",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=50,
            fg_color=VERDE,
            hover_color=VERDE_ESC,
            corner_radius=10,
            command=self._iniciar,
        )
        self.btn_converter.grid(row=0, column=0, sticky="ew")

        # ── Progresso ─────────────────────────────────────────────────────────
        prog_frame = ctk.CTkFrame(self.scroll, fg_color=FUNDO_CARD,
                                  corner_radius=12)
        prog_frame.grid(row=row, column=0, sticky="ew", padx=24, pady=(4, 4))
        prog_frame.grid_columnconfigure(0, weight=1)
        row += 1

        prog_header = ctk.CTkFrame(prog_frame, fg_color="transparent")
        prog_header.grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 4))
        prog_header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            prog_header,
            text="Progresso",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=TEXTO_SUB,
        ).grid(row=0, column=0, sticky="w")

        self.lbl_pct = ctk.CTkLabel(
            prog_header,
            text="0%",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=VERDE,
        )
        self.lbl_pct.grid(row=0, column=1, sticky="e")

        self.barra = ctk.CTkProgressBar(
            prog_frame, height=10,
            fg_color=FUNDO_INPUT,
            progress_color=VERDE,
            corner_radius=5,
        )
        self.barra.set(0)
        self.barra.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))

        self.lbl_status = ctk.CTkLabel(
            prog_frame,
            text="Aguardando...",
            font=ctk.CTkFont(size=10),
            text_color=TEXTO_SUB,
        )
        self.lbl_status.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 10))

        # ── Log ───────────────────────────────────────────────────────────────
        log_frame = ctk.CTkFrame(self.scroll, fg_color=FUNDO_CARD,
                                 corner_radius=12)
        log_frame.grid(row=row, column=0, sticky="ew", padx=24, pady=(0, 20))
        log_frame.grid_columnconfigure(0, weight=1)
        row += 1

        log_top = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_top.grid(row=0, column=0, sticky="ew", padx=16, pady=(10, 4))
        log_top.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            log_top,
            text="Log de Processamento",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=TEXTO_SUB,
        ).grid(row=0, column=0, sticky="w")

        ctk.CTkButton(
            log_top, text="Limpar", width=60, height=22,
            fg_color=BORDA, hover_color=FUNDO_INPUT,
            font=ctk.CTkFont(size=10),
            command=self._limpar_log,
        ).grid(row=0, column=1, sticky="e")

        self.log_box = ctk.CTkTextbox(
            log_frame,
            height=180,
            fg_color=FUNDO_INPUT,
            text_color=TEXTO,
            font=ctk.CTkFont(family="Courier New", size=10),
            corner_radius=8,
            border_width=1,
            border_color=BORDA,
            wrap="none",
        )
        self.log_box.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 6))

        ctk.CTkLabel(
            log_frame,
            text=f"📄  Arquivo de log salvo em: {LOG_FILE}",
            font=ctk.CTkFont(size=10),
            text_color=TEXTO_SUB,
        ).grid(row=2, column=0, sticky="w", padx=16, pady=(0, 12))

    # ── Helpers de layout ─────────────────────────────────────────────────────

    def _card(self, row: int, titulo: str, builder_fn) -> int:
        card = ctk.CTkFrame(self.scroll, fg_color=FUNDO_CARD, corner_radius=12)
        card.grid(row=row, column=0, sticky="ew", padx=24, pady=(8, 0))
        card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            card, text=titulo,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=TEXTO,
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(12, 6))

        builder_fn(card)
        return row + 1

    def _pasta_row(self, parent, label: str, var: ctk.StringVar,
                   comando, grid_row: int):
        ctk.CTkLabel(
            parent, text=label,
            font=ctk.CTkFont(size=11),
            text_color=TEXTO_SUB,
            width=110, anchor="w",
        ).grid(row=grid_row, column=0, sticky="w", padx=(16, 4), pady=4)

        entry = ctk.CTkEntry(
            parent, textvariable=var,
            fg_color=FUNDO_INPUT, border_color=BORDA,
            text_color=TEXTO, placeholder_text="Selecione uma pasta...",
            height=36, font=ctk.CTkFont(size=11),
        )
        entry.grid(row=grid_row, column=1, sticky="ew", padx=4, pady=4)

        ctk.CTkButton(
            parent, text="Procurar", width=90, height=36,
            fg_color=BORDA, hover_color=VERDE_ESC,
            font=ctk.CTkFont(size=11),
            command=comando,
        ).grid(row=grid_row, column=2, padx=(4, 16), pady=4)

    # ── Seções ────────────────────────────────────────────────────────────────

    def _build_pastas(self, parent):
        parent.grid_columnconfigure(1, weight=1)
        self._pasta_row(parent, "Pasta de Entrada:", self.pasta_entrada,
                        self._sel_entrada, 1)
        self._pasta_row(parent, "Pasta de Saída:",   self.pasta_saida,
                        self._sel_saida,   2)

        ctk.CTkLabel(
            parent,
            text="ℹ  Se a Pasta de Saída ficar vazia, os PDFs/Excel serão gerados na mesma pasta dos XMLs.",
            font=ctk.CTkFont(size=10), text_color=TEXTO_SUB,
        ).grid(row=3, column=0, columnspan=3, sticky="w", padx=16, pady=(0, 10))

    def _build_formato(self, parent):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))

        # PDF
        self.chk_pdf = ctk.CTkCheckBox(
            f, text="PDF  (DANFE / DACTE)",
            variable=self.gerar_pdf,
            checkbox_width=20, checkbox_height=20,
            fg_color=VERDE, hover_color=VERDE_ESC,
            font=ctk.CTkFont(size=12),
            text_color=TEXTO,
        )
        self.chk_pdf.grid(row=0, column=0, padx=(0, 24), pady=4)

        # Excel
        excel_estado = "normal" if EXCEL_OK else "disabled"
        self.chk_excel = ctk.CTkCheckBox(
            f, text="Excel  (.xlsx)  — resumo e produtos",
            variable=self.gerar_excel,
            state=excel_estado,
            checkbox_width=20, checkbox_height=20,
            fg_color=AZUL, hover_color="#3A7FCC",
            font=ctk.CTkFont(size=12),
            text_color=TEXTO if EXCEL_OK else TEXTO_SUB,
        )
        self.chk_excel.grid(row=0, column=1, padx=0, pady=4)

        if not EXCEL_OK:
            ctk.CTkLabel(
                parent, text="⚠  openpyxl não instalado: pip install openpyxl",
                font=ctk.CTkFont(size=10), text_color=AVISO,
            ).grid(row=2, column=0, sticky="w", padx=16, pady=(0, 8))

    def _build_opcoes(self, parent):
        parent.grid_columnconfigure(0, weight=1)

        # ── Posição do recibo ─────────────────────────────────────────────────
        row_recibo = ctk.CTkFrame(parent, fg_color="transparent")
        row_recibo.grid(row=1, column=0, sticky="w", padx=16, pady=(0, 8))

        ctk.CTkLabel(
            row_recibo, text="Posição do Recibo (DANFE):",
            font=ctk.CTkFont(size=11), text_color=TEXTO_SUB,
        ).grid(row=0, column=0, padx=(0, 12))

        self.seg_recibo = ctk.CTkSegmentedButton(
            row_recibo,
            values=["Topo", "Rodapé"],
            command=lambda v: self.recibo_pos.set(
                "topo" if v == "Topo" else "baixo"),
            fg_color=FUNDO_INPUT,
            selected_color=VERDE, selected_hover_color=VERDE_ESC,
            unselected_color=FUNDO_INPUT, unselected_hover_color=BORDA,
            font=ctk.CTkFont(size=11),
            width=160, height=32,
        )
        self.seg_recibo.set("Topo")
        self.seg_recibo.grid(row=0, column=1)

        # ── Arquivos existentes ───────────────────────────────────────────────
        row_sobr = ctk.CTkFrame(parent, fg_color="transparent")
        row_sobr.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 12))

        ctk.CTkLabel(
            row_sobr, text="Arquivos já convertidos:",
            font=ctk.CTkFont(size=11), text_color=TEXTO_SUB,
        ).grid(row=0, column=0, padx=(0, 12))

        self.seg_sobr = ctk.CTkSegmentedButton(
            row_sobr,
            values=["Pular", "Sobrescrever"],
            command=lambda v: self.sobrescrever.set(v == "Sobrescrever"),
            fg_color=FUNDO_INPUT,
            selected_color=VERDE, selected_hover_color=VERDE_ESC,
            unselected_color=FUNDO_INPUT, unselected_hover_color=BORDA,
            font=ctk.CTkFont(size=11),
            width=200, height=32,
        )
        self.seg_sobr.set("Pular")
        self.seg_sobr.grid(row=0, column=1)

    # ── Seletores de pasta ────────────────────────────────────────────────────

    def _sel_entrada(self):
        d = filedialog.askdirectory(title="Selecione a pasta com os XMLs")
        if d:
            self.pasta_entrada.set(d)
            if not self.pasta_saida.get():
                self.pasta_saida.set(d)

    def _sel_saida(self):
        d = filedialog.askdirectory(title="Selecione a pasta de destino")
        if d:
            self.pasta_saida.set(d)

    # ── Log ───────────────────────────────────────────────────────────────────

    def _log(self, msg: str, nivel: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        linha = f"[{ts}]  {msg}\n"

        # Escreve na caixa da interface
        self.log_box.configure(state="normal")
        self.log_box.insert("end", linha)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

        # Escreve no arquivo de log
        nivel_fn = getattr(logger, nivel, logger.info)
        nivel_fn(msg)

    def _limpar_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    # ── Progresso ─────────────────────────────────────────────────────────────

    def _set_progresso(self, valor: float, status: str = ""):
        self.barra.set(valor)
        self.lbl_pct.configure(text=f"{int(valor * 100)}%")
        if status:
            self.lbl_status.configure(text=status)

    # ── Conversão ─────────────────────────────────────────────────────────────

    def _iniciar(self):
        if self.processando:
            return

        # Validações
        entrada = self.pasta_entrada.get().strip()
        if not entrada or not Path(entrada).is_dir():
            messagebox.showerror("Erro", "Selecione uma pasta de entrada válida.")
            return
        if not self.gerar_pdf.get() and not self.gerar_excel.get():
            messagebox.showerror("Erro", "Selecione ao menos um formato de saída.")
            return
        if self.gerar_pdf.get() and not FISCAL_OK:
            messagebox.showerror(
                "Erro",
                "brazilfiscalreport não está instalado.\n"
                "Execute: pip install brazilfiscalreport qrcode[pil] pillow",
            )
            return

        self.processando = True
        self.btn_converter.configure(
            text="⏳  Processando...", state="disabled", fg_color=BORDA)

        t = threading.Thread(target=self._processar, daemon=True)
        t.start()

    def _processar(self):
        try:
            self._processar_interno()
        except Exception as e:
            self._log(f"ERRO GERAL: {e}", ERRO)
        finally:
            self.processando = False
            self.btn_converter.configure(
                text="▶   Iniciar Conversão",
                state="normal", fg_color=VERDE,
            )

    def _processar_interno(self):
        entrada  = Path(self.pasta_entrada.get().strip())
        saida_s  = self.pasta_saida.get().strip()
        saida    = Path(saida_s) if saida_s else entrada
        saida.mkdir(parents=True, exist_ok=True)

        # Separador de sessão no arquivo de log
        logger.info("=" * 60)
        logger.info(f"NOVA SESSÃO  —  entrada: {entrada}  |  saída: {saida}")
        logger.info("=" * 60)

        # Config PDF
        if FISCAL_OK and self.gerar_pdf.get():
            pos = ReceiptPosition.TOP if self.recibo_pos.get() == "topo" \
                  else ReceiptPosition.BOTTOM
            danfe_cfg = DanfeConfig(receipt_pos=pos)
            dacte_cfg = DacteConfig()
        else:
            danfe_cfg = dacte_cfg = None

        arquivos = sorted(entrada.rglob("*.xml"))
        total    = len(arquivos)

        if total == 0:
            self._log("Nenhum arquivo .xml encontrado na pasta selecionada.")
            self._set_progresso(0, "Nenhum XML encontrado.")
            return

        self._log(f"Encontrados {total} arquivo(s) XML.")
        self._set_progresso(0, f"0 / {total}")

        ok = erro = pulado = 0
        regs_nfe: list[dict] = []
        regs_cte: list[dict] = []

        for i, xml_path in enumerate(arquivos, 1):
            rel      = xml_path.relative_to(entrada)
            tipo     = detectar_tipo(xml_path)
            nome     = xml_path.name

            frac = i / total
            self._set_progresso(frac, f"{i} / {total}  —  {nome}")

            if tipo == "desconhecido":
                self._log(f"⚠  {nome}  →  tipo não reconhecido, ignorado.", "warning")
                pulado += 1
                continue

            # Extração para Excel
            if self.gerar_excel.get() and EXCEL_OK:
                try:
                    if tipo == "nfe":
                        regs_nfe.append(extrair_nfe(xml_path))
                    else:
                        regs_cte.append(extrair_cte(xml_path))
                except Exception as e:
                    self._log(f"⚠  {nome}  →  erro ao extrair dados: {e}", "warning")

            # Geração PDF
            if self.gerar_pdf.get() and FISCAL_OK:
                pdf_path = (saida / rel).with_suffix(".pdf")
                pdf_path.parent.mkdir(parents=True, exist_ok=True)

                if pdf_path.exists() and not self.sobrescrever.get():
                    self._log(f"⏭  {nome}  →  PDF já existe, pulado.")
                    pulado += 1
                    continue

                try:
                    converter_pdf(xml_path, pdf_path, tipo, danfe_cfg, dacte_cfg)
                    kb = pdf_path.stat().st_size / 1024
                    self._log(f"✅  {nome}  →  {pdf_path.name}  ({kb:.1f} KB)")
                    ok += 1
                except Exception as e:
                    self._log(f"❌  {nome}  →  {e}", "error")
                    erro += 1
            else:
                if tipo in ("nfe", "cte"):
                    ok += 1

        # Gera Excel consolidado
        if self.gerar_excel.get() and EXCEL_OK and (regs_nfe or regs_cte):
            ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_xl  = f"notas_fiscais_{ts_str}.xlsx"
            caminho_xl = saida / nome_xl
            try:
                self._log(f"📊  Gerando Excel: {nome_xl} ...")
                gerar_excel(regs_nfe, regs_cte, caminho_xl)
                kb = caminho_xl.stat().st_size / 1024
                self._log(f"✅  Excel gerado: {nome_xl}  ({kb:.1f} KB)")
            except Exception as e:
                self._log(f"❌  Erro ao gerar Excel: {e}", "error")

        # Resumo final
        self._set_progresso(1.0, "Concluído!")
        self._log("─" * 50)
        self._log(
            f"Concluído  —  ✅ {ok} convertidos  "
            f"⏭ {pulado} pulados  ❌ {erro} erros"
        )
        self._log(f"📁  Saída: {saida.resolve()}")

        messagebox.showinfo(
            "Concluído",
            f"Processamento finalizado!\n\n"
            f"✅  Convertidos : {ok}\n"
            f"⏭  Pulados     : {pulado}\n"
            f"❌  Erros       : {erro}\n\n"
            f"📁  {saida.resolve()}",
        )


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
